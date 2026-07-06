# -*- coding: utf-8 -*-
"""Genera todos los archivos del Kit Paletas de WhatsApp."""
import csv
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from content import (
    RECETAS, COMBINACIONES, MENSAJES, PLAN_7_DIAS,
    CHECKLIST_VENTA, CHECKLIST_PRODUCCION, ERRORES_COMUNES, FOTOS_TIPS,
    TECNICAS_BASICAS, COMBOS_SUGERIDOS, EJEMPLO_PRECIO, ALERGENOS_RECETA,
)

ROOT = Path(__file__).resolve().parent.parent
PRODUTO = ROOT / "produto"
BUILD = Path(__file__).resolve().parent
STYLES = (BUILD / "styles.css").read_text(encoding="utf-8")

ROSA = "FF4F8B"
CREMA = "FFF4E6"
CHOCOLATE = "5C2E1F"


def tipo_badge(tipo):
    m = {"Frutal": "badge-frutal", "Cremosa": "badge-cremosa", "Rellena": "badge-rellena", "Estilo postre": "badge-postre"}
    return m.get(tipo, "badge-frutal")


def receta_html(r):
    ing = "".join(f"<li>{i}</li>" for i in r["ingredientes"])
    pasos = "".join(f"<li>{p}</li>" for p in r["pasos"])
    dif = "badge-facil" if r["dificultad"] == "Fácil" else "badge-media"
    alergia = ALERGENOS_RECETA.get(r["num"])
    alergia_html = f'<p class="alergeno-tag"><strong>Contiene:</strong> {alergia}</p>' if alergia else ""
    consejo = r.get("consejo", r.get("dica", ""))
    return f"""
    <div class="receta-card">
      <div class="receta-header">
        <div class="receta-num">{r['num']}</div>
        <div class="receta-title"><h3>{r['nombre']}</h3></div>
      </div>
      <div class="receta-meta">
        <span class="badge {tipo_badge(r['tipo'])}">{r['tipo']}</span>
        <span class="badge {dif}">{r['dificultad']}</span>
      </div>
      <p><strong>Preparación:</strong> {r['prep']} &nbsp;|&nbsp; <strong>Congelación:</strong> {r['congelacion']} &nbsp;|&nbsp; <strong>Rinde:</strong> {r['rendimiento']}</p>
      {alergia_html}
      <div class="receta-grid">
        <div class="ingredientes"><h4>Ingredientes</h4><ul>{ing}</ul></div>
        <div><h4>Preparación</h4><ol>{pasos}</ol></div>
        <div class="col-full dica-vender"><strong>💡 Consejo para vender:</strong> {consejo}</div>
      </div>
      <p class="precio-nota">💰 Calcula el precio con la calculadora según tus costos locales. Ajusta dulzor, tamaño y empaque según tu público.</p>
    </div>"""


CATEGORIAS_RECETAS = [
    (1, "Paletas frutales", "Frescas, coloridas y económicas. Ideales para empezar.", 1, 7),
    (8, "Paletas cremosas", "Clásicas y cremosas. Fáciles de vender a todo público.", 8, 14),
    (15, "Paletas rellenas", "Más valor percibido. Requieren un paso extra.", 15, 21),
    (22, "Paletas estilo postre", "Sabores diferenciados para destacar en tu menú.", 22, 30),
]


def recetas_por_categoria():
    pages = []
    for _, titulo, desc, start, end in CATEGORIAS_RECETAS:
        subset = [r for r in RECETAS if start <= r["num"] <= end]
        cards = "".join(receta_html(r) for r in subset[:2])
        pages.append(f"""<div class="page category-page">
          {page_header()}
          <span class="section-tag">Recetas</span>
          <h2>{titulo}</h2>
          <p>{desc}</p>
          {cards}
        </div>""")
        remaining = subset[2:]
        for i in range(0, len(remaining), 2):
            batch = remaining[i:i + 2]
            cards = "".join(receta_html(r) for r in batch)
            pages.append(f'<div class="page">{page_header()}{cards}</div>')
    return "".join(pages)


def page_header():
    return '<div class="page-header"><strong>Paletas de WhatsApp</strong><span>Kit completo</span></div>'


def build_kit_html():
    combos = "".join(
        f'<li><strong>{c["nombre"]}</strong> — {c["porque"]}</li>' for c in COMBINACIONES
    )

    mensajes_kit = ""
    for cat, items in [
        ("Anunciar menú", MENSAJES["menu"][:5]),
        ("Estado / Story", MENSAJES["status"][:5]),
        ("Promoción", MENSAJES["promo"][:5]),
    ]:
        for m in items:
            mensajes_kit += f'<div class="mensaje"><div class="mensaje-cat">{cat}</div>{m}</div>'

    faq_kit = "".join(f'<div class="mensaje"><div class="mensaje-cat">Pregunta frecuente</div>{m}</div>' for m in MENSAJES["faq"][:5])

    plan_html = ""
    for d in PLAN_7_DIAS:
        tareas = "".join(f"<li>{t}</li>" for t in d["tareas"])
        plan_html += f'''<div class="dia-card"><span class="dia-num">Día {d["dia"]}</span>
          <h3>{d["titulo"]}</h3>
          <p class="dia-meta"><strong>Tiempo:</strong> {d["duracion"]} &nbsp;|&nbsp; <strong>Meta:</strong> {d["meta"]}</p>
          <ul>{tareas}</ul></div>'''

    tecnicas = ""
    for t in TECNICAS_BASICAS:
        pasos = "".join(f"<li>{p}</li>" for p in t["pasos"])
        tecnicas += f"<h3>{t['titulo']}</h3><ol>{pasos}</ol>"

    combos_html = ""
    for c in COMBOS_SUGERIDOS:
        combos_html += f'<div class="box"><h4>{c["nombre"]}</h4><p>{c["contenido"]}</p><p class="precio-nota">{c["nota"]}</p></div>'

    ep = EJEMPLO_PRECIO
    errores = "".join(f"<li>{e}</li>" for e in ERRORES_COMUNES)
    fotos = "".join(f"<li>{f}</li>" for f in FOTOS_TIPS)
    checklist = "".join(f'<div class="checklist-item"><div class="check-box"></div><span>{c}</span></div>' for c in CHECKLIST_VENTA)
    recetas_pages = recetas_por_categoria()

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Kit Paletas de WhatsApp</title>
<style>{STYLES}</style>
</head>
<body>

<div class="page cover">
  <div class="cover-badge">Kit Digital</div>
  <div class="cover-icon">🍭</div>
  <h1>Paletas de WhatsApp</h1>
  <p class="subtitle">Recetas, precios y mensajes listos para empezar desde casa.</p>
  <div class="tagline">Prepara. Calcula. Publica.</div>
  <div class="method">
    <span>🥄 Preparar</span>
    <span>💰 Precio</span>
    <span>📱 Publicar</span>
  </div>
  <div class="cover-footer">Método 3P — Recetas + Precios + Mensajes</div>
</div>

<div class="page">
  {page_header()}
  <h2>Índice</h2>
  <div class="toc-item"><span>Bienvenida</span><span>3</span></div>
  <div class="toc-item"><span>Cómo usar este kit</span><span>4</span></div>
  <div class="toc-item"><span>El Método 3P</span><span>5</span></div>
  <div class="toc-item"><span>Antes de empezar</span><span>6</span></div>
  <div class="toc-item"><span>Higiene y seguridad</span><span>7</span></div>
  <div class="toc-item"><span>Técnicas básicas</span><span>8</span></div>
  <div class="toc-item"><span>30 Recetas de paletas</span><span>9+</span></div>
  <div class="toc-item"><span>Combinaciones que llaman la atención</span><span>—</span></div>
  <div class="toc-item"><span>Cómo elegir tus primeras recetas</span><span>—</span></div>
  <div class="toc-item"><span>Cómo calcular precios + ejemplo</span><span>—</span></div>
  <div class="toc-item"><span>Combos y precios</span><span>—</span></div>
  <div class="toc-item"><span>Cómo usar la calculadora</span><span>—</span></div>
  <div class="toc-item"><span>Menú editable para WhatsApp</span><span>—</span></div>
  <div class="toc-item"><span>Mensajes y preguntas frecuentes</span><span>—</span></div>
  <div class="toc-item"><span>Plan de 7 días</span><span>—</span></div>
  <div class="toc-item"><span>Fotos, errores y checklist</span><span>—</span></div>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Inicio</span>
  <h2>Bienvenida al Kit Paletas de WhatsApp</h2>
  <p>Felicidades por acceder al kit. Este material fue creado para ayudarte a organizar una pequeña oferta de paletas caseras desde casa, sin complicarte y sin cobrar "a ojo".</p>
  <div class="box box-rosa">
    <p><strong>Importante:</strong> Este kit no promete ingresos garantizados ni resultados automáticos. Te entrega materiales prácticos — recetas, herramientas de precio, menú y mensajes — para que organices tu oferta con más claridad.</p>
  </div>
  <p>La idea es simple: muchas personas quieren vender algo desde casa pero no saben por dónde empezar. Este kit te guía con el <strong>Método 3P</strong>: Preparar, Precio y Publicar.</p>
  <h3>Lo que incluye tu kit</h3>
  <ul>
    <li>30 recetas de paletas caseras, simples y realistas</li>
    <li>Calculadora de precios en Excel (también compatible con Google Sheets)</li>
    <li>Menú editable para publicar en WhatsApp</li>
    <li>60+ mensajes listos para copiar y adaptar</li>
    <li>Plan de 7 días paso a paso</li>
    <li>Checklist de compras y producción</li>
  </ul>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Guía</span>
  <h2>Cómo usar este kit</h2>
  <ol class="steps">
    <li><strong>Elige 3 a 5 recetas</strong> para empezar. No intentes hacer las 30 de una vez.</li>
    <li><strong>Calcula tus costos</strong> con la calculadora incluida. Ajusta a los precios de tu ciudad.</li>
    <li><strong>Arma tu menú</strong> con los sabores y precios que definiste.</li>
    <li><strong>Publica tus mensajes</strong> en WhatsApp y empieza a tomar pedidos con más claridad.</li>
  </ol>
  <div class="box">
    <p>💡 <strong>Consejo:</strong> Imprime o guarda el checklist al final. Repásalo antes de cada tanda de producción.</p>
  </div>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Método</span>
  <h2>El Método 3P</h2>
  <div class="box box-menta">
    <h3>🥄 Preparar</h3>
    <p>Elige recetas simples y visuales. Empieza con sabores que conozcas y que sean fáciles de conseguir en tu zona.</p>
  </div>
  <div class="box">
    <h3>💰 Precio</h3>
    <p>Calcula ingredientes, porciones, empaque y margen estimado. No cobres "a ojo" — usa la calculadora para tener una referencia real.</p>
  </div>
  <div class="box box-rosa">
    <h3>📱 Publicar</h3>
    <p>Usa un menú claro y mensajes listos para ofrecer tus paletas. WhatsApp es tu vitrina: sé claro, amable y directo con sabores y precios.</p>
  </div>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Preparación</span>
  <h2>Antes de empezar</h2>
  <h3>Materiales básicos</h3>
  <ul>
    <li>Moldes para paletas</li>
    <li>Palitos</li>
    <li>Bolsitas o empaques</li>
    <li>Etiquetas simples</li>
    <li>Licuadora</li>
    <li>Vasos medidores</li>
    <li>Recipientes</li>
    <li>Congelador</li>
    <li>Frutas</li>
    <li>Leche o base cremosa</li>
    <li>Chocolate o coberturas, si aplica</li>
  </ul>
  <div class="box-aviso">
    ⚠️ Los ingredientes y costos pueden variar según tu país, ciudad y proveedores. Ajusta siempre los precios a tu realidad.
  </div>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Seguridad</span>
  <h2>Reglas básicas de higiene y seguridad</h2>
  <p>No necesitas un laboratorio. Solo buenos hábitos:</p>
  <ul>
    <li>Lava tus manos y utensilios antes de preparar</li>
    <li>Usa ingredientes frescos y revisa fechas de vencimiento</li>
    <li>Mantén la cadena de frío: congela a tiempo y transporta con hieleras si es necesario</li>
    <li>Etiqueta los sabores, especialmente si tienes varios</li>
    <li>Informa al cliente si contiene leche, maní, nueces, gluten u otros alérgenos</li>
    <li>Revisa las normas locales si vas a vender alimentos en tu ciudad</li>
  </ul>
  <div class="box-aviso">
    Esta es orientación general, no asesoría legal. Cada municipio puede tener reglas distintas para venta de alimentos caseros.
  </div>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Técnica</span>
  <h2>Técnicas básicas antes de preparar</h2>
  <p>Estos tres puntos resuelven la mayoría de problemas cuando empiezas:</p>
  {tecnicas}
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Recetas</span>
  <h2>30 Recetas de Paletas Caseras</h2>
  <p>Organizadas por tipo. Empieza con 3 a 5 sabores y ve agregando según la demanda.</p>
</div>

{recetas_pages}

<div class="page">
  {page_header()}
  <span class="section-tag">Ideas</span>
  <h2>Combinaciones que suelen llamar la atención</h2>
  <ul class="combo-list">{combos}</ul>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Estrategia</span>
  <h2>Cómo elegir tus primeras recetas</h2>
  <p>No empieces con 30 sabores. Empieza con <strong>3 a 5</strong>:</p>
  <ul>
    <li>1 frutal (fácil, económica, colorida)</li>
    <li>1 cremosa (clásica, todos la entienden)</li>
    <li>1 rellena (diferenciadora, más valor percibido)</li>
    <li>1 infantil/colorida (para familias)</li>
    <li>1 premium (para cobrar un poco más)</li>
  </ul>
  <div class="box box-menta">
    <h3>Ejemplo — Menú Semana 1</h3>
    <ul>
      <li>Fresa con crema</li>
      <li>Mango con limón</li>
      <li>Chocolate con galleta</li>
      <li>Coco tropical</li>
      <li>Cheesecake de fresa</li>
    </ul>
  </div>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Precios</span>
  <h2>Cómo calcular precios</h2>
  <p>Para saber si una paleta vale la pena, considera:</p>
  <ul>
    <li>Ingredientes</li>
    <li>Empaque (bolsa, etiqueta)</li>
    <li>Palito</li>
    <li>Energía o gas, si quieres estimar</li>
    <li>Pérdidas (lo que se derrama o no sale bien)</li>
    <li>Margen que deseas</li>
  </ul>
  <div class="box">
    <h3>Fórmula simple</h3>
    <p><strong>Costo total de producción ÷ cantidad de paletas = costo por paleta</strong></p>
    <p><strong>Precio sugerido = costo por paleta ÷ (1 − margen%)</strong></p>
  </div>
  <h3>Ejemplo con números ficticios</h3>
  <table>
    <tr><th>Concepto</th><th>Valor</th></tr>
    <tr><td>Costo ingredientes (8 paletas)</td><td>${ep['ingredientes']:.2f}</td></tr>
    <tr><td>Empaque + palitos</td><td>${ep['empaque']:.2f}</td></tr>
    <tr><td>Costo por paleta</td><td>${ep['costo_unitario']:.2f}</td></tr>
    <tr><td>Margen deseado</td><td>{ep['margen']}%</td></tr>
    <tr><td><strong>Precio sugerido</strong></td><td><strong>${ep['precio_sugerido']:.2f}</strong></td></tr>
  </table>
  <p class="precio-nota">{ep['nota']}</p>
  <div class="box-aviso">
    La margen es una estimación, no una garantía de ganancia. Los costos cambian y la demanda varía. Revisa tus precios cada semana si es necesario.
  </div>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Herramienta</span>
  <h2>Cómo usar la calculadora</h2>
  <p>Abre el archivo <strong>Calculadora_Precios_Paletas.xlsx</strong> (o súbelo a Google Sheets):</p>
  <ol class="steps">
    <li>Escribe el nombre de la receta</li>
    <li>Agrega cada ingrediente con su costo de paquete y cantidad usada</li>
    <li>Coloca el costo de empaque y otros gastos</li>
    <li>Escribe cuántas paletas produjiste</li>
    <li>Mira el costo aproximado por paleta</li>
    <li>Define tu margen deseado (%)</li>
    <li>Revisa el precio sugerido y la ganancia estimada</li>
  </ol>
  <p>También tiene pestañas para lista de compras, menú y pedidos.</p>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Combos</span>
  <h2>Cómo armar combos sin perder margen</h2>
  <p>Los combos ayudan a vender más unidades por pedido. La regla: nunca armes un combo sin calcular antes.</p>
  {combos_html}
  <div class="box-aviso">Tip: un combo con 10–15% de descuento sobre el precio individual suele ser atractivo sin regalar tu trabajo.</div>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Menú</span>
  <h2>Menú editable para WhatsApp</h2>
  <p>Usa el archivo <strong>Menu_Editable_Paletas</strong> o copia estos modelos:</p>
  <div class="menu-modelo">
    <h3>🍭 Paletas Caseras de la Semana</h3>
    <div class="menu-linea"><span>Fresa con crema</span><span class="precio">$<span class="campo-editable">&nbsp;&nbsp;&nbsp;&nbsp;</span></span></div>
    <div class="menu-linea"><span>Mango con limón</span><span class="precio">$<span class="campo-editable">&nbsp;&nbsp;&nbsp;&nbsp;</span></span></div>
    <div class="menu-linea"><span>Chocolate cremoso</span><span class="precio">$<span class="campo-editable">&nbsp;&nbsp;&nbsp;&nbsp;</span></span></div>
    <p style="text-align:center;margin-top:14px;font-size:10pt;">📱 Pedidos por WhatsApp — Disponible hasta agotar stock</p>
  </div>
  <div class="menu-modelo">
    <h3>🌈 Menú Dulce de Hoy</h3>
    <p><strong>Frutales:</strong> <span class="campo-editable">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span></p>
    <p><strong>Cremosas:</strong> <span class="campo-editable">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span></p>
    <p><strong>Rellenas:</strong> <span class="campo-editable">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span></p>
    <p><strong>Combo promo:</strong> 3 paletas por $<span class="campo-editable">&nbsp;&nbsp;&nbsp;&nbsp;</span></p>
  </div>
  <div class="menu-modelo">
    <h3>✨ Paletas Artesanales</h3>
    <p><strong>Sabores especiales:</strong> <span class="campo-editable">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span></p>
    <p><strong>Combo familiar (6 u.):</strong> $<span class="campo-editable">&nbsp;&nbsp;&nbsp;&nbsp;</span></p>
    <p><strong>Pedidos con anticipación:</strong> 24h mínimo</p>
  </div>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Mensajes</span>
  <h2>Mensajes listos para WhatsApp</h2>
  <p>Selección de mensajes del pack completo. Copia, adapta y personaliza:</p>
  {mensajes_kit}
  <p style="margin-top:12px;font-size:10pt;">📄 El pack completo con 60+ mensajes está en <strong>Mensajes_para_Vender_Paletas.pdf</strong></p>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">FAQ</span>
  <h2>Preguntas frecuentes de clientes</h2>
  <p>Cuando te escriban por primera vez, estas respuestas te ahorran tiempo. Personaliza los campos entre [corchetes]:</p>
  {faq_kit}
  <p style="margin-top:12px;font-size:10pt;">Más respuestas en el archivo de mensajes.</p>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Plan</span>
  <h2>Plan de 7 días</h2>
  {plan_html}
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Fotos</span>
  <h2>Fotos simples que venden mejor</h2>
  <ul>{fotos}</ul>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Evita</span>
  <h2>Errores comunes</h2>
  <ol>{errores}</ol>
</div>

<div class="page">
  {page_header()}
  <span class="section-tag">Checklist</span>
  <h2>Checklist final — Antes de vender</h2>
  {checklist}
</div>

</body>
</html>"""
    return html


def build_mensajes_html():
    sections = [
        ("A) Mensajes para anunciar menú", MENSAJES["menu"]),
        ("B) Mensajes de estado / story", MENSAJES["status"]),
        ("C) Mensajes de promoción de fin de semana", MENSAJES["promo"]),
        ("D) Mensajes para clientes antiguos", MENSAJES["clientes"]),
        ("E) Mensajes de últimas unidades", MENSAJES["ultimas"]),
        ("F) Mensajes para pedidos", MENSAJES["pedidos"]),
        ("G) Respuestas a preguntas frecuentes", MENSAJES["faq"]),
    ]
    body = ""
    for title, items in sections:
        body += f'<div class="page">{page_header()}<h2>{title}</h2>'
        for i, m in enumerate(items, 1):
            body += f'<div class="mensaje"><strong>{i}.</strong> {m}</div>'
        body += "</div>"
    return f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Mensajes para Vender Paletas</title><style>{STYLES}</style></head><body>
<div class="page cover"><div class="cover-icon">💬</div><h1>Mensajes para Vender Paletas</h1><p class="subtitle">60 textos listos para copiar en WhatsApp</p></div>
<div class="page">{page_header()}
  <h2>Cómo usar estos mensajes</h2>
  <p>Copia el texto que necesites y reemplaza lo que está entre <strong>[corchetes]</strong> con tu información real (sabores, precios, zona).</p>
  <div class="box-aviso">No prometas lo que no puedes cumplir. Ajusta cantidades y plazos a lo que realmente puedes producir.</div>
</div>
{body}
</body></html>"""


def build_plan_html():
    plan = ""
    for d in PLAN_7_DIAS:
        tareas = "".join(f"<li>{t}</li>" for t in d["tareas"])
        plan += f'''<div class="dia-card"><span class="dia-num">Día {d["dia"]}</span>
          <h3>{d["titulo"]}</h3>
          <p class="dia-meta"><strong>Tiempo:</strong> {d["duracion"]} &nbsp;|&nbsp; <strong>Meta del día:</strong> {d["meta"]}</p>
          <ul>{tareas}</ul></div>'''
    return f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Plan 7 Días Paletas</title><style>{STYLES}</style></head><body>
<div class="page cover"><div class="cover-icon">📅</div><h1>Plan de 7 Días</h1><p class="subtitle">Tu guía paso a paso para empezar a vender paletas</p></div>
<div class="page">{page_header()}<h2>Tu semana de inicio</h2><p>Sigue un día a la vez. No te saltes pasos.</p>{plan}</div>
</body></html>"""


def build_checklist_html():
    items = "".join(f'<div class="checklist-item"><div class="check-box"></div><span>{c}</span></div>' for c in CHECKLIST_VENTA)
    prod_html = "".join(f'<div class="checklist-item"><div class="check-box"></div><span>{p}</span></div>' for p in CHECKLIST_PRODUCCION)
    return f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Checklist Paletas</title><style>{STYLES}</style></head><body>
<div class="page cover"><div class="cover-icon">✅</div><h1>Checklist Paletas</h1><p class="subtitle">Compras, producción y venta</p></div>
<div class="page">{page_header()}<h2>Antes de vender</h2>{items}</div>
<div class="page">{page_header()}<h2>Día de producción</h2>{prod_html}</div>
<div class="page">{page_header()}<h2>Lista de compras — Ingredientes</h2>
<ul><li>Leche, crema, leche condensada</li><li>Azúcar, miel, vainilla</li><li>Frutas frescas o congeladas</li><li>Cacao, chocolate, yogur</li><li>Mermelada, dulce de leche</li><li>Queso crema (para postres)</li></ul>
<h3>Materiales</h3>
<ul><li>Moldes, palitos, bolsas, etiquetas</li><li>Bandejas para transporte</li></ul>
</div>
</body></html>"""


def build_menu_html():
    return f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Menú Editable Paletas</title><style>
{STYLES}
.menu-edit {{ max-width: 500px; margin: 0 auto; }}
input, textarea {{ border: none; border-bottom: 2px dashed var(--rosa); background: transparent; font-family: inherit; font-size: inherit; color: var(--chocolate); width: 100%; padding: 4px; }}
input:focus, textarea:focus {{ outline: 2px solid var(--rosa); border-radius: 4px; }}
.menu-linea {{ display: flex; gap: 8px; align-items: center; margin-bottom: 8px; }}
.btn-copy {{
  background: var(--rosa); color: white; border: none; border-radius: 12px;
  padding: 12px 20px; font-family: inherit; font-weight: 700; font-size: 13px;
  cursor: pointer; width: 100%; margin-top: 12px;
}}
.btn-copy:hover {{ opacity: 0.9; }}
.preview-box {{
  background: #E8F5E9; border-radius: 12px; padding: 14px; margin-top: 16px;
  font-size: 13px; white-space: pre-wrap; display: none;
}}
.preview-box.visible {{ display: block; }}
@media print {{ input, textarea {{ border-bottom: 1px solid #ccc; }} .btn-copy {{ display: none; }} }}
</style></head><body>

<div class="page cover"><div class="cover-icon">📋</div><h1>Menú Editable</h1><p class="subtitle">Completa los campos, genera el texto y publícalo en WhatsApp</p></div>

<div class="page menu-edit">
  <div class="menu-modelo" id="menu-simple">
    <h3>🍭 Paletas Caseras de la Semana</h3>
    <div class="menu-linea"><input class="sabor" placeholder="Sabor 1" value="Fresa con crema"><input class="precio" placeholder="$" style="width:70px;text-align:right" value=""></div>
    <div class="menu-linea"><input class="sabor" placeholder="Sabor 2" value="Mango con limón"><input class="precio" placeholder="$" style="width:70px;text-align:right" value=""></div>
    <div class="menu-linea"><input class="sabor" placeholder="Sabor 3" value="Chocolate cremoso"><input class="precio" placeholder="$" style="width:70px;text-align:right" value=""></div>
    <div class="menu-linea"><input class="sabor" placeholder="Sabor 4" value=""><input class="precio" placeholder="$" style="width:70px;text-align:right" value=""></div>
    <div class="menu-linea"><input class="sabor" placeholder="Sabor 5" value=""><input class="precio" placeholder="$" style="width:70px;text-align:right" value=""></div>
    <textarea id="nota-menu" rows="2" style="margin-top:12px">📱 Pedidos por WhatsApp — Disponible hasta agotar stock</textarea>
    <button type="button" class="btn-copy" onclick="generarMenu()">Copiar menú para WhatsApp</button>
    <div class="preview-box" id="preview"></div>
  </div>
</div>

<div class="page menu-edit">
  <div class="menu-modelo" style="background:#FFF0F5">
    <h3>🌈 Menú Dulce de Hoy</h3>
    <p><strong>Frutales:</strong> <input id="frutales" placeholder="ej: mango, sandía"></p>
    <p><strong>Cremosas:</strong> <input id="cremosas" placeholder="ej: vainilla, chocolate"></p>
    <p><strong>Rellenas:</strong> <input id="rellenas" placeholder="ej: dulce de leche"></p>
    <p><strong>Combo promo:</strong> 3 paletas por $<input id="combo-precio" style="width:80px;display:inline" placeholder="0.00"></p>
    <button type="button" class="btn-copy" onclick="generarColorido()">Copiar menú colorido</button>
    <div class="preview-box" id="preview2"></div>
  </div>
</div>

<div class="page menu-edit">
  <div class="menu-modelo" style="background:var(--menta)">
    <h3>✨ Paletas Artesanales</h3>
    <p><strong>Especiales de la semana:</strong></p>
    <textarea id="especiales" rows="3" placeholder="Tus sabores premium..."></textarea>
    <p><strong>Combo familiar (6 u.):</strong> $<input id="familiar" style="width:80px;display:inline"></p>
    <p><strong>Anticipación:</strong> <input id="anticipacion" value="24 horas mínimo"></p>
    <button type="button" class="btn-copy" onclick="generarPremium()">Copiar menú premium</button>
    <div class="preview-box" id="preview3"></div>
  </div>
</div>

<div class="page">
  <h2>Textos para estado / story</h2>
  <div class="mensaje">🍭 Paletas caseras disponibles hoy. Escríbeme para ver sabores y precios.</div>
  <div class="mensaje">Menú de la semana en mi estado anterior 👆</div>
  <div class="mensaje">¿Cuál te provoca? Pregunta por disponibilidad</div>
</div>

<script>
function copiar(texto, previewId) {{
  navigator.clipboard.writeText(texto).then(() => {{
    const el = document.getElementById(previewId);
    el.textContent = '✅ Copiado! Pega en WhatsApp:\\n\\n' + texto;
    el.classList.add('visible');
  }}).catch(() => {{
    const el = document.getElementById(previewId);
    el.textContent = texto;
    el.classList.add('visible');
  }});
}}
function generarMenu() {{
  const sabores = document.querySelectorAll('#menu-simple .menu-linea');
  let lineas = ['🍭 *Paletas Caseras de la Semana*', ''];
  sabores.forEach(row => {{
    const s = row.querySelector('.sabor').value.trim();
    const p = row.querySelector('.precio').value.trim();
    if (s) lineas.push(p ? `• ${{s}} — $${{p}}` : `• ${{s}}`);
  }});
  lineas.push('', document.getElementById('nota-menu').value.trim());
  copiar(lineas.join('\\n'), 'preview');
}}
function generarColorido() {{
  const t = [
    '🌈 *Menú Dulce de Hoy*', '',
    `🍓 Frutales: ${{document.getElementById('frutales').value || '...'}}`,
    `🍦 Cremosas: ${{document.getElementById('cremosas').value || '...'}}`,
    `🍫 Rellenas: ${{document.getElementById('rellenas').value || '...'}}`,
    `🎁 Combo: 3 paletas por $${{document.getElementById('combo-precio').value || '...'}}`, '',
    '📱 Escríbeme para apartar las tuyas'
  ];
  copiar(t.join('\\n'), 'preview2');
}}
function generarPremium() {{
  const t = [
    '✨ *Paletas Artesanales*', '',
    `Sabores especiales:\\n${{document.getElementById('especiales').value || '...'}}`, '',
    `👨‍👩‍👧 Combo familiar (6 u.): $${{document.getElementById('familiar').value || '...'}}`,
    `⏰ Pedidos con ${{document.getElementById('anticipacion').value}} de anticipación`, '',
    '📱 Reserva por WhatsApp'
  ];
  copiar(t.join('\\n'), 'preview3');
}}
</script>
</body></html>"""


def build_entrega_html():
    return """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Tu Kit Paletas de WhatsApp</title>
<link href="https://fonts.googleapis.com/css2?family=Nunito:wght@400;600;700;800&display=swap" rel="stylesheet">
<style>
:root {
  --rosa: #FF4F8B; --naranja: #FF7A1A; --crema: #FFF4E6;
  --chocolate: #5C2E1F; --menta: #CFF7E2;
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: 'Nunito', system-ui, sans-serif;
  background: linear-gradient(160deg, var(--crema), #FFE8F0 50%, var(--menta));
  min-height: 100vh; color: var(--chocolate);
}
.container { max-width: 720px; margin: 0 auto; padding: 32px 20px; }
header { text-align: center; margin-bottom: 32px; }
header .icon { font-size: 56px; }
header h1 { font-size: 28px; font-weight: 800; margin: 12px 0 8px; }
header p { color: #6B5E57; font-size: 15px; }
.badge {
  display: inline-block; background: var(--rosa); color: white;
  font-size: 11px; font-weight: 700; letter-spacing: 0.1em;
  text-transform: uppercase; padding: 5px 14px; border-radius: 50px; margin-bottom: 16px;
}
.card {
  background: white; border-radius: 18px; padding: 20px 24px;
  margin-bottom: 14px; box-shadow: 0 4px 20px rgba(92,46,31,0.08);
  display: flex; align-items: center; gap: 16px; text-decoration: none; color: inherit;
  transition: transform 0.15s, box-shadow 0.15s;
}
.card:hover { transform: translateY(-2px); box-shadow: 0 8px 28px rgba(92,46,31,0.12); }
.card-icon { font-size: 32px; flex-shrink: 0; }
.card-info h3 { font-size: 15px; font-weight: 700; margin-bottom: 2px; }
.card-info p { font-size: 12px; color: #6B5E57; }
.card-arrow { margin-left: auto; color: var(--rosa); font-weight: 800; font-size: 18px; }
.section-title {
  font-size: 12px; font-weight: 700; text-transform: uppercase;
  letter-spacing: 0.08em; color: #6B5E57; margin: 24px 0 12px;
}
.steps {
  background: white; border-radius: 18px; padding: 20px 24px;
  box-shadow: 0 4px 20px rgba(92,46,31,0.08);
}
.steps ol { margin-left: 20px; }
.steps li { margin-bottom: 8px; font-size: 14px; }
footer { text-align: center; margin-top: 32px; font-size: 12px; color: #6B5E57; }
footer a { color: var(--rosa); }
</style>
</head>
<body>
<div class="container">
  <header>
    <div class="badge">Acceso confirmado</div>
    <div class="icon">🍭</div>
    <h1>¡Tu kit está listo!</h1>
    <p>Paletas de WhatsApp — Recetas, precios y mensajes listos para empezar desde casa.</p>
  </header>

  <div class="section-title">Empieza por aquí (en orden)</div>
  <div class="steps" style="margin-bottom:20px">
    <ol>
      <li><strong>Kit Principal</strong> — lee "Cómo usar este kit" (10 min)</li>
      <li><strong>Calculadora</strong> — pon precios de tu ciudad</li>
      <li><strong>Menú Editable</strong> — elige 3 a 5 sabores y copia el texto</li>
      <li><strong>Plan de 7 Días</strong> — un paso por día</li>
    </ol>
  </div>

  <div class="section-title">Descarga tus archivos</div>

  <a class="card" href="produto/Kit_Paletas_de_WhatsApp.pdf" download>
    <div class="card-icon">📘</div>
    <div class="card-info"><h3>Kit Principal (PDF)</h3><p>30 recetas + guía completa</p></div>
    <span class="card-arrow">↓</span>
  </a>
  <a class="card" href="produto/Calculadora_Precios_Paletas.xlsx" download>
    <div class="card-icon">📊</div>
    <div class="card-info"><h3>Calculadora de Precios</h3><p>Excel — también funciona en Google Sheets</p></div>
    <span class="card-arrow">↓</span>
  </a>
  <a class="card" href="produto/Menu_Editable_Paletas.html" target="_blank">
    <div class="card-icon">📋</div>
    <div class="card-info"><h3>Menú Editable</h3><p>Completa y publica en WhatsApp</p></div>
    <span class="card-arrow">→</span>
  </a>
  <a class="card" href="produto/Mensajes_para_Vender_Paletas.pdf" download>
    <div class="card-icon">💬</div>
    <div class="card-info"><h3>Mensajes para WhatsApp</h3><p>60 textos listos para copiar</p></div>
    <span class="card-arrow">↓</span>
  </a>
  <a class="card" href="produto/Plan_7_Dias_Paletas.pdf" download>
    <div class="card-icon">📅</div>
    <div class="card-info"><h3>Plan de 7 Días</h3><p>Tu guía paso a paso</p></div>
    <span class="card-arrow">↓</span>
  </a>
  <a class="card" href="produto/Checklist_Paletas.pdf" download>
    <div class="card-icon">✅</div>
    <div class="card-info"><h3>Checklist</h3><p>Compras, producción y venta</p></div>
    <span class="card-arrow">↓</span>
  </a>

  <div class="section-title">Tips rápidos</div>
  <div class="steps">
    <ul style="margin-left:20px">
      <li><strong>Calculadora en Google Sheets:</strong> sube el archivo .xlsx a Google Drive → Abrir con Hojas de cálculo</li>
      <li><strong>Menú:</strong> usa el botón "Copiar menú" y pega directo en WhatsApp</li>
      <li><strong>Empieza con 3 sabores</strong>, no con 30</li>
    </ul>
  </div>

  <div class="section-title">Ruta de la primera semana</div>
  <div class="steps">
    <ol>
      <li>Descarga el <strong>Kit Principal</strong> y léelo hasta la sección "Cómo usar este kit"</li>
      <li>Abre la <strong>Calculadora</strong> y pon los precios de tu ciudad</li>
      <li>Elige <strong>3 a 5 recetas</strong> y completa el <strong>Menú Editable</strong></li>
      <li>Sigue el <strong>Plan de 7 Días</strong> — un paso por día</li>
      <li>Usa los <strong>Mensajes</strong> para publicar en WhatsApp</li>
    </ol>
  </div>

  <footer>
    <p>Prepara. Calcula. Publica. 🍭</p>
    <p style="margin-top:8px">¿Dudas? Escríbenos por WhatsApp desde tu correo de compra.</p>
  </footer>
</div>
</body>
</html>"""


def style_header(ws, title, fill_color=ROSA):
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=fill_color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def build_calculator():
    wb = Workbook()
    thin = Side(style="thin", color="E8DFD6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Aba 0: Inicio (instrucciones) ---
    ws0 = wb.create_sheet("Inicio", 0)
    ws0.merge_cells("A1:D1")
    ws0["A1"] = "Cómo usar esta calculadora"
    ws0["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws0["A1"].fill = PatternFill("solid", fgColor=ROSA)
    ws0.row_dimensions[1].height = 28
    instrucciones = [
        "1. Ve a la pestaña 'Calculadora' y escribe el nombre de tu receta.",
        "2. Agrega cada ingrediente con el costo del paquete y cuánto usaste.",
        "3. Abajo, pon el costo de empaque, palito y otros gastos.",
        "4. Escribe cuántas paletas produjiste.",
        "5. Mira el precio sugerido según tu margen.",
        "6. Usa 'Lista de compras', 'Menú' y 'Pedidos' para organizarte.",
        "",
        "Tip: Sube este archivo a Google Drive → clic derecho → Abrir con Hojas de cálculo.",
        "Los números del ejemplo en 'Ejemplo' son ficticios.",
    ]
    for i, line in enumerate(instrucciones, 3):
        ws0.cell(row=i, column=1, value=line)
    ws0.column_dimensions["A"].width = 70

    # --- Aba 1: Calculadora ---
    ws = wb.active
    ws.title = "Calculadora"
    style_header(ws, "Calculadora de Precios — Paletas de WhatsApp")

    headers = ["Ingrediente", "Costo del paquete", "Cantidad total del paquete", "Cantidad usada", "Costo usado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=CHOCOLATE)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 20

    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    ws["A2"] = "Nombre de la receta:"
    ws["B2"] = "Paleta de fresa cremosa"
    ws["B2"].font = Font(bold=True)

    sample_ings = [
        ("Fresas", 3.50, 500, 200),
        ("Leche", 1.20, 1000, 250),
        ("Crema", 2.00, 200, 100),
        ("Azúcar", 1.50, 1000, 30),
    ]
    for i, (ing, costo, total, usado) in enumerate(sample_ings, 4):
        ws.cell(row=i, column=1, value=ing)
        ws.cell(row=i, column=2, value=costo)
        ws.cell(row=i, column=3, value=total)
        ws.cell(row=i, column=4, value=usado)
        ws.cell(row=i, column=5, value=f"=IF(OR(C{i}=0,D{i}=0),0,B{i}/C{i}*D{i})")

    for row in range(4, 14):
        if not ws.cell(row=row, column=1).value:
            ws.cell(row=row, column=5, value=f"=IF(OR(C{row}=0,D{row}=0),0,B{row}/C{row}*D{row})")

    ws["A15"] = "— Costos fijos por tanda —"
    ws["A15"].font = Font(bold=True, italic=True)
    ws["A16"] = "Empaque (costo por unidad):"
    ws["B16"] = 0.15
    ws["A17"] = "Palito (costo por unidad):"
    ws["B17"] = 0.05
    ws["A18"] = "Otros costos (gas, etiquetas, etc.):"
    ws["B18"] = 0.50
    ws["A19"] = "Cantidad de paletas producidas:"
    ws["B19"] = 8
    ws["A20"] = "Costo total ingredientes:"
    ws["B20"] = "=SUM(E4:E13)"
    ws["A21"] = "Costo empaques (unidad × cantidad):"
    ws["B21"] = "=(B16+B17)*B19+B18"
    ws["A22"] = "Costo total de producción:"
    ws["B22"] = "=B20+B21"
    ws["A23"] = "Costo por paleta:"
    ws["B23"] = "=IF(B19=0,0,B22/B19)"
    ws["A24"] = "Margen deseado %:"
    ws["B24"] = 40
    ws["A25"] = "Precio sugerido:"
    ws["B25"] = "=IF(B24>=100,0,B23/(1-B24/100))"
    ws["A26"] = "Ganancia estimada por unidad:"
    ws["B26"] = "=B25-B23"

    for r in range(16, 27):
        ws.cell(row=r, column=1).font = Font(bold=True)
    for r in range(20, 27):
        ws.cell(row=r, column=2).number_format = '"$"#,##0.00'
    for row in range(4, 14):
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border

    ws["A28"] = "Las celdas en verde se calculan solas. Solo edita ingredientes y costos."
    ws.merge_cells("A28:E28")
    ws["A28"].font = Font(italic=True, color="6B5E57", size=9)

    # --- Aba 2: Ejemplo ---
    ws2 = wb.create_sheet("Ejemplo")
    style_header(ws2, "Ejemplo ilustrativo — Paleta de fresa cremosa", "FF7A1A")
    ws2["A3"] = "⚠️ Ejemplo ilustrativo. Ajusta con tus costos reales."
    ws2.merge_cells("A3:F3")

    ex_headers = ["Ingrediente", "Costo paquete", "Cant. total", "Cant. usada", "Costo usado"]
    for col, h in enumerate(ex_headers, 1):
        c = ws2.cell(row=5, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=CHOCOLATE)

    ejemplo = [
        ("Fresas", 3.50, 500, 200),
        ("Leche", 1.20, 1000, 250),
        ("Crema", 2.00, 200, 100),
        ("Azúcar", 1.50, 1000, 30),
        ("Vainilla", 2.00, 100, 5),
        ("Empaque", 0.15, 1, 1),
        ("Palito", 0.05, 1, 1),
    ]
    for i, (ing, costo, total, usado) in enumerate(ejemplo, 6):
        ws2.cell(row=i, column=1, value=ing)
        ws2.cell(row=i, column=2, value=costo)
        ws2.cell(row=i, column=3, value=total)
        ws2.cell(row=i, column=4, value=usado)
        ws2.cell(row=i, column=5, value=f"=B{i}/C{i}*D{i}")

    ws2["A14"] = "Paletas producidas:"
    ws2["B14"] = 8
    ws2["A15"] = "Costo total:"
    ws2["B15"] = "=SUM(E6:E12)"
    ws2["A16"] = "Costo por paleta:"
    ws2["B16"] = "=B15/B14"
    ws2["A17"] = "Margen 40%:"
    ws2["B17"] = "=B16/(1-0.4)"
    ws2["A18"] = "Ganancia/unidad:"
    ws2["B18"] = "=B17-B16"

    # --- Aba 3: Lista de compras ---
    ws3 = wb.create_sheet("Lista de compras")
    style_header(ws3, "Lista de Compras")
    for col, h in enumerate(["Ingrediente", "Cantidad", "Proveedor", "Precio", "Observaciones"], 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=CHOCOLATE)
        ws3.column_dimensions[get_column_letter(col)].width = 20
    defaults = ["Leche", "Crema", "Fresas", "Azúcar", "Moldes", "Palitos", "Bolsas"]
    for i, d in enumerate(defaults, 4):
        ws3.cell(row=i, column=1, value=d)

    # --- Aba 4: Menú ---
    ws4 = wb.create_sheet("Menú")
    style_header(ws4, "Mi Menú de Paletas")
    for col, h in enumerate(["Sabor", "Costo estimado", "Precio", "Disponible", "Observaciones"], 1):
        c = ws4.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=CHOCOLATE)

    # --- Aba 5: Pedidos ---
    ws5 = wb.create_sheet("Pedidos")
    style_header(ws5, "Control de Pedidos")
    for col, h in enumerate(["Cliente", "Teléfono", "Sabor", "Cantidad", "Total", "Pagado", "Entregado", "Observaciones"], 1):
        c = ws5.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=CHOCOLATE)
        ws5.column_dimensions[get_column_letter(col)].width = 16

    xlsx_path = PRODUTO / "Calculadora_Precios_Paletas.xlsx"
    wb.save(xlsx_path)
    print(f"  OK {xlsx_path.name}")

    # CSV versions
    csv_dir = PRODUTO
    with open(csv_dir / "Calculadora_Precios_Paletas.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Calculadora de Precios — Paletas de WhatsApp"])
        w.writerow(["Ingrediente", "Costo del paquete", "Cantidad total", "Cantidad usada", "Costo usado"])
        w.writerow(["", "", "", "", "=(B/C)*D"])
        w.writerow([])
        w.writerow(["Cantidad paletas", "8"])
        w.writerow(["Margen %", "40"])
        w.writerow(["Fórmula costo/paleta", "=SUMA(costos_usados)/cantidad_paletas"])
        w.writerow(["Fórmula precio", "=costo_paleta/(1-margen/100)"])
    print(f"  OK Calculadora_Precios_Paletas.csv")


async def html_to_pdf(html_path, pdf_path):
    from playwright.async_api import async_playwright
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        page = await browser.new_page()
        await page.goto(html_path.as_uri(), wait_until="networkidle")
        await page.pdf(
            path=str(pdf_path),
            format="A4",
            print_background=True,
            margin={"top": "0", "right": "0", "bottom": "0", "left": "0"},
        )
        await browser.close()


def generate_pdfs():
    import asyncio
    pairs = [
        ("Kit_Paletas_de_WhatsApp.html", "Kit_Paletas_de_WhatsApp.pdf"),
        ("Mensajes_para_Vender_Paletas.html", "Mensajes_para_Vender_Paletas.pdf"),
        ("Plan_7_Dias_Paletas.html", "Plan_7_Dias_Paletas.pdf"),
        ("Checklist_Paletas.html", "Checklist_Paletas.pdf"),
        ("Menu_Editable_Paletas.html", "Menu_Editable_Paletas.pdf"),
    ]
    for html_name, pdf_name in pairs:
        html_path = PRODUTO / html_name
        pdf_path = PRODUTO / pdf_name
        if html_path.exists():
            asyncio.run(html_to_pdf(html_path, pdf_path))
            print(f"  OK {pdf_name}")


def main():
    PRODUTO.mkdir(parents=True, exist_ok=True)
    print("Generando Kit Paletas de WhatsApp...\n")

    files = {
        "Kit_Paletas_de_WhatsApp.html": build_kit_html(),
        "Mensajes_para_Vender_Paletas.html": build_mensajes_html(),
        "Plan_7_Dias_Paletas.html": build_plan_html(),
        "Checklist_Paletas.html": build_checklist_html(),
        "Menu_Editable_Paletas.html": build_menu_html(),
    }
    for name, content in files.items():
        path = PRODUTO / name
        path.write_text(content, encoding="utf-8")
        print(f"  OK {name}")

    (ROOT / "entrega.html").write_text(build_entrega_html(), encoding="utf-8")
    print(f"  OK entrega.html")

    print("\nGenerando calculadora...")
    build_calculator()

    print("\nGenerando PDFs (puede tardar un momento)...")
    try:
        generate_pdfs()
    except Exception as e:
        print(f"  WARN Error generando PDFs: {e}")
        print("  Puedes exportar manualmente desde los HTML (Ctrl+P → Guardar como PDF)")

    readme = """# Paletas de WhatsApp — Kit Digital

Kit digital completo para vender paletas caseras por WhatsApp.

## Contenido

| Archivo | Descripción |
|---------|-------------|
| `produto/Kit_Paletas_de_WhatsApp.pdf` | PDF principal con 30 recetas y guía completa |
| `produto/Calculadora_Precios_Paletas.xlsx` | Calculadora de precios (5 pestañas) |
| `produto/Menu_Editable_Paletas.html` | Menú editable para WhatsApp |
| `produto/Mensajes_para_Vender_Paletas.pdf` | 50+ mensajes listos |
| `produto/Plan_7_Dias_Paletas.pdf` | Plan paso a paso |
| `produto/Checklist_Paletas.pdf` | Checklist imprimible |
| `entrega.html` | Página de descarga para área de miembros |

## Regenerar archivos

```bash
cd paletas-de-whatsapp/build
pip install openpyxl playwright
python -m playwright install chromium
python build.py
```

## Exportar PDF manualmente

Si los PDFs no se generaron automáticamente:
1. Abre cualquier `.html` en Chrome
2. Ctrl+P → Destino: Guardar como PDF
3. Activa "Gráficos de fondo"

## Usar la calculadora

1. Abre `Calculadora_Precios_Paletas.xlsx` en Excel o súbelo a Google Sheets
2. Pestaña **Calculadora**: ingresa ingredientes y costos
3. Pestaña **Ejemplo**: referencia con números ficticios
4. Pestañas **Lista de compras**, **Menú** y **Pedidos**: organiza tu operación

## Área de miembros

Publica `entrega.html` en tu servidor o Vercel. Los enlaces apuntan a `produto/`.

## Editar recetas

Las recetas están en `build/content.py`. Después de editar, ejecuta `python build.py` para regenerar.
"""
    (ROOT / "README.md").write_text(readme, encoding="utf-8")
    print(f"  OK README.md")
    print("\nKit generado en paletas-de-whatsapp/")
    sync_public()


def sync_public():
    """Copia archivos a public/ para deploy en Vercel."""
    import shutil
    public = ROOT.parent / "public" / "paletas-de-whatsapp"
    public_prod = public / "produto"
    public_prod.mkdir(parents=True, exist_ok=True)
    for f in PRODUTO.iterdir():
        if f.is_file():
            shutil.copy2(f, public_prod / f.name)
    shutil.copy2(ROOT / "entrega.html", public / "entrega.html")
    print("  OK public/paletas-de-whatsapp/ sincronizado")


if __name__ == "__main__":
    main()
